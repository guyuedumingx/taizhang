"""多维度台账统计查询服务。

查询链路（见《特殊栏位处理与填写格式提示方案》3.3）：
  ① SQL 粗筛：template/status/team/时间范围/关键词（纯关系列，两库高效）
  ② Python 清洗层：按 field_normalizers.json 配置归一化特殊栏位
  ③ 内存过滤：字段级筛选（数值类走清洗后值）
  ④ 内存聚合：SUM/COUNT + 数据质量报告
  ⑤ 排序、内存分页

聚合放 Python 层的理由：中文单位换算/大写金额等复杂规则只有应用层能干净实现，
顺带规避 Oracle JSON 函数版本兼容问题，质量报告几乎免费。
"""
import datetime as _dt
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException
from sqlalchemy import or_
from sqlalchemy.orm import Session, selectinload

from app import models
from app.schemas.statistics_query import (
    DataQualityReport,
    FieldFilterCondition,
    FieldQuality,
    LedgerQueryItem,
    LedgerQueryResponse,
    QueryField,
    StatisticsQueryRequest,
    SuspiciousItem,
    SystemFilters,
)
from app.services.normalizers import NormalizeResult, find_pipeline, get_template_pipelines, run_pipeline

SORTABLE_SYSTEM_FIELDS = {"id", "name", "status", "approval_status", "created_at", "updated_at"}
MAX_EXPORT_ROWS = 10000
MAX_SUSPICIOUS_ITEMS = 50

_SYSTEM_COLUMNS = ["ID", "台账名称", "模板", "状态", "审批状态", "团队", "创建人", "创建时间"]


def _fmt_number(v: float) -> str:
    return f"{v:g}"


def _parse_range(range_list: Optional[List[str]]) -> Tuple[Optional[_dt.datetime], Optional[_dt.datetime]]:
    """["2026-01-01", "2026-05-27"] → (起始零点, 结束日次日零点)。非法输入静默忽略。"""
    if not range_list or len(range_list) != 2:
        return None, None
    start = end = None
    try:
        if range_list[0]:
            start = _dt.datetime.fromisoformat(range_list[0])
        if range_list[1]:
            end = _dt.datetime.fromisoformat(range_list[1])
            if len(range_list[1]) <= 10:  # 纯日期，含当日
                end = end + _dt.timedelta(days=1)
    except ValueError:
        return None, None
    return start, end


class StatisticsService:

    # ---------- 查询主流程 ----------

    @staticmethod
    def query_ledgers(db: Session, query: StatisticsQueryRequest, fetch_all: bool = False) -> LedgerQueryResponse:
        ledgers = StatisticsService._sql_prefilter(db, query)
        pipelines_by_template = StatisticsService._pipelines_for(ledgers)

        nr_cache: Dict[Tuple[int, str], Optional[NormalizeResult]] = {}

        def pipeline_result(led: models.Ledger, field_name: str) -> Optional[NormalizeResult]:
            key = (led.id, field_name)
            if key not in nr_cache:
                entry = pipelines_by_template.get(led.template.name if led.template else "", {}).get(field_name)
                nr_cache[key] = (
                    run_pipeline((led.data or {}).get(field_name), entry.get("pipeline"), entry.get("invalid_policy"))
                    if entry else None
                )
            return nr_cache[key]

        # ②③ 字段级过滤（数值条件走清洗后值）
        matched: List[models.Ledger] = []
        for led in ledgers:
            ok = True
            for field_name, cond in (query.field_filters or {}).items():
                entry = pipelines_by_template.get(led.template.name if led.template else "", {}).get(field_name)
                nr = pipeline_result(led, field_name) if entry else None
                if not StatisticsService._match_condition((led.data or {}).get(field_name), nr, cond):
                    ok = False
                    break
            if ok:
                matched.append(led)

        # ④ 数据质量聚合（对过滤后的结果集）
        data_quality = StatisticsService._aggregate_quality(matched, pipelines_by_template, pipeline_result)

        # 排序
        matched = StatisticsService._sort(matched, query, pipelines_by_template, pipeline_result)

        total = len(matched)
        if fetch_all:
            page_items = matched[:MAX_EXPORT_ROWS]
        else:
            page = max(query.page, 1)
            page_size = max(min(query.page_size, 200), 1)
            page_items = matched[(page - 1) * page_size: page * page_size]

        return LedgerQueryResponse(
            items=[StatisticsService._to_item(led) for led in page_items],
            total=total,
            page=query.page if not fetch_all else 1,
            page_size=query.page_size if not fetch_all else total,
            data_quality=data_quality,
        )

    # ---------- SQL 粗筛 ----------

    @staticmethod
    def _sql_prefilter(db: Session, query: StatisticsQueryRequest) -> List[models.Ledger]:
        q = db.query(models.Ledger).options(
            selectinload(models.Ledger.team),
            selectinload(models.Ledger.template),
            selectinload(models.Ledger.creator),
        )
        if query.template_ids:
            q = q.filter(models.Ledger.template_id.in_(query.template_ids))

        sf: SystemFilters = query.system_filters or SystemFilters()
        if sf.status:
            q = q.filter(models.Ledger.status.in_(sf.status))
        if sf.approval_status:
            q = q.filter(models.Ledger.approval_status.in_(sf.approval_status))
        if sf.team_ids:
            q = q.filter(models.Ledger.team_id.in_(sf.team_ids))
        if sf.created_by_ids:
            q = q.filter(models.Ledger.created_by_id.in_(sf.created_by_ids))

        start, end = _parse_range(sf.created_at_range)
        if start:
            q = q.filter(models.Ledger.created_at >= start)
        if end:
            q = q.filter(models.Ledger.created_at < end)
        start, end = _parse_range(sf.updated_at_range)
        if start:
            q = q.filter(models.Ledger.updated_at >= start)
        if end:
            q = q.filter(models.Ledger.updated_at < end)

        if query.keyword and query.keyword.strip():
            kw = f"%{query.keyword.strip()}%"
            q = q.filter(or_(models.Ledger.name.ilike(kw), models.Ledger.description.ilike(kw)))

        return q.all()

    @staticmethod
    def _pipelines_for(ledgers: List[models.Ledger]) -> Dict[str, Dict[str, dict]]:
        """结果集中出现的每个模板 → {字段名: 清洗配置}。"""
        result: Dict[str, Dict[str, dict]] = {}
        for led in ledgers:
            tn = led.template.name if led.template else ""
            if tn not in result:
                result[tn] = get_template_pipelines(tn)
        return result

    # ---------- 条件匹配 ----------

    @staticmethod
    def _match_condition(raw: Any, nr: Optional[NormalizeResult], cond: FieldFilterCondition) -> bool:
        op = cond.operator
        value = cond.value
        if op == "contains":
            return value is not None and str(value) in str(raw if raw is not None else "")
        if op == "in":
            candidates = value if isinstance(value, list) else [value]
            return raw in candidates or str(raw) in [str(c) for c in candidates]
        if op == "equals":
            if nr is not None and nr.is_numeric and not nr.is_suspicious:
                try:
                    return nr.value == float(value)
                except (TypeError, ValueError):
                    return False
            return raw == value

        # 数值比较类：配置了清洗规则的必须用清洗后值（normalized_between 即文档 3.5 的补充操作符）
        uses_pipeline = nr is not None
        if op in ("gte", "lte", "between", "normalized_between"):
            if uses_pipeline:
                if not nr.is_numeric or nr.is_suspicious:
                    return False
                v = nr.value
            else:
                try:
                    v = float(str(raw))
                except (TypeError, ValueError):
                    return False
            try:
                if op == "gte":
                    return v >= float(value)
                if op == "lte":
                    return v <= float(value)
                lo, hi = float(value[0]), float(value[1])
                return lo <= v <= hi
            except (TypeError, ValueError, IndexError):
                return False
        return True

    # ---------- 质量聚合 ----------

    @staticmethod
    def _aggregate_quality(ledgers, pipelines_by_template, pipeline_result) -> DataQualityReport:
        acc: Dict[Tuple[str, str], FieldQuality] = {}
        total = len(ledgers)
        for led in ledgers:
            tn = led.template.name if led.template else ""
            for field_name in pipelines_by_template.get(tn, {}):
                nr = pipeline_result(led, field_name)
                if nr is None:
                    continue
                key = (tn, field_name)
                fq = acc.setdefault(key, FieldQuality(field_name=field_name))
                if nr.is_suspicious:
                    fq.suspicious_count += 1
                    if len(fq.suspicious_items) < MAX_SUSPICIOUS_ITEMS:
                        fq.suspicious_items.append(SuspiciousItem(
                            ledger_id=led.id,
                            ledger_name=led.name,
                            field=field_name,
                            raw=nr.raw,
                            reason=nr.note or "无法解析",
                        ))
                elif nr.is_numeric:
                    fq.numeric_count += 1
                    fq.sum = (fq.sum or 0) + nr.value
                    # 原始值与清洗值不同 → 计为被清洗修复
                    if str(nr.raw).strip() != _fmt_number(nr.value):
                        fq.cleaned_count += 1

        # 同名字段跨模板合并（统计口径按字段名）
        merged: Dict[str, FieldQuality] = {}
        for (tn, field_name), fq in acc.items():
            target = merged.setdefault(field_name, FieldQuality(field_name=field_name))
            target.sum = (target.sum or 0) + (fq.sum or 0)
            target.numeric_count += fq.numeric_count
            target.cleaned_count += fq.cleaned_count
            target.suspicious_count += fq.suspicious_count
            remaining = MAX_SUSPICIOUS_ITEMS - len(target.suspicious_items)
            target.suspicious_items.extend(fq.suspicious_items[:max(remaining, 0)])
        return DataQualityReport(total_count=total, fields=list(merged.values()))

    # ---------- 排序与转换 ----------

    @staticmethod
    def _sort(ledgers, query: StatisticsQueryRequest, pipelines_by_template, pipeline_result):
        reverse = (query.sort_order or "desc").lower() != "asc"
        sort_by = query.sort_by or "created_at"

        if sort_by in SORTABLE_SYSTEM_FIELDS:
            def sys_key(led):
                v = getattr(led, sort_by)
                return (v is None, v)
            return sorted(ledgers, key=sys_key, reverse=reverse)

        def field_key(led):
            raw = (led.data or {}).get(sort_by)
            nr = pipeline_result(led, sort_by)
            if nr is not None:
                if nr.is_numeric and not nr.is_suspicious:
                    return (0, nr.value, "")
                return (1, 0.0, nr.raw)
            try:
                return (0, float(str(raw)), "")
            except (TypeError, ValueError):
                return (1, 0.0, str(raw if raw is not None else ""))
        return sorted(ledgers, key=field_key, reverse=reverse)

    @staticmethod
    def _to_item(led: models.Ledger) -> LedgerQueryItem:
        return LedgerQueryItem(
            id=led.id,
            name=led.name,
            template_id=led.template_id,
            template_name=led.template.name if led.template else None,
            status=led.status,
            approval_status=led.approval_status,
            team_id=led.team_id,
            team_name=led.team.name if led.team else None,
            created_by_id=led.created_by_id,
            created_by_name=led.creator.name if led.creator else None,
            created_at=led.created_at,
            updated_at=led.updated_at,
            data=led.data or {},
        )

    # ---------- 字段元信息 ----------

    @staticmethod
    def get_query_fields(db: Session, template_id: int) -> List[QueryField]:
        template = db.query(models.Template).filter(models.Template.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="模板不存在")
        fields = (
            db.query(models.Field)
            .filter(models.Field.template_id == template_id)
            .order_by(models.Field.order)
            .all()
        )
        result = []
        for f in fields:
            options = f.options if isinstance(f.options, list) else None
            result.append(QueryField(
                name=f.name or "",
                label=f.label or f.name or "",
                type=f.type,
                required=bool(f.required),
                options=options,
                has_pipeline=find_pipeline(template.name, f.name or "") is not None,
            ))
        return result

    # ---------- 导出 ----------

    @staticmethod
    def export_query_results(db: Session, query: StatisticsQueryRequest) -> Tuple[BytesIO, str, str]:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        response = StatisticsService.query_ledgers(db, query, fetch_all=True)
        pipelines_by_template = {
            tn: get_template_pipelines(tn)
            for tn in {it.template_name or "" for it in response.items}
        }

        wb = Workbook()
        ws = wb.active
        ws.title = "查询结果"
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="EEEEEE")

        # 列结构：单模板时附带其字段列（按模板字段顺序）
        field_columns: List[Tuple[str, str]] = []  # (field_name, label)
        template_ids = {it.template_id for it in response.items}
        if len(template_ids) == 1 and None not in template_ids:
            fields = (
                db.query(models.Field)
                .filter(models.Field.template_id == template_ids.pop())
                .order_by(models.Field.order)
                .all()
            )
            field_columns = [(f.name or "", f.label or f.name or "") for f in fields]

        header = _SYSTEM_COLUMNS + [label for _, label in field_columns]
        for col, title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill

        suspicious_rows = []
        for r, it in enumerate(response.items, 2):
            values = [it.id, it.name, it.template_name, it.status, it.approval_status,
                      it.team_name, it.created_by_name,
                      it.created_at.strftime("%Y-%m-%d %H:%M") if it.created_at else ""]
            for col, v in enumerate(values, 1):
                ws.cell(row=r, column=col, value=v)
            for c, (fname, _) in enumerate(field_columns, len(_SYSTEM_COLUMNS) + 1):
                raw = it.data.get(fname)
                entry = pipelines_by_template.get(it.template_name or "", {}).get(fname)
                if entry:
                    nr = run_pipeline(raw, entry.get("pipeline"), entry.get("invalid_policy"))
                    if nr.is_suspicious:
                        ws.cell(row=r, column=c, value=f"⚠ {nr.raw}")
                        suspicious_rows.append([it.id, it.name, fname, nr.raw, nr.note or "无法解析"])
                        continue
                    if nr.is_numeric:
                        ws.cell(row=r, column=c, value=nr.value)
                        continue
                ws.cell(row=r, column=c, value=raw if raw is not None else "")

        ws2 = wb.create_sheet("可疑数据")
        for col, title in enumerate(["台账ID", "台账名称", "字段", "原始值", "原因"], 1):
            cell = ws2.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
        for r, row in enumerate(suspicious_rows, 2):
            for col, v in enumerate(row, 1):
                ws2.cell(row=r, column=col, value=v)

        for col in range(1, len(header) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        for col in range(1, 6):
            ws2.column_dimensions[get_column_letter(col)].width = 22

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = _dt.datetime.now().strftime("%Y%m%d%H%M")
        filename = f"台账汇总查询_{ts}.xlsx"
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return buf, filename, content_type
